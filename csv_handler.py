import csv
import logging
import time
import re
from openpyxl import load_workbook
import pandas as pd


logging.basicConfig(filename='logs.txt', filemode='a',
                    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.Formatter.converter = time.gmtime
logger = logging.getLogger(__name__)


def gtin_set(csv_file_path):
    unique_gtins = set()
    with open(csv_file_path) as file:
        rows = file.readlines()
        for row in rows:
            result = find_datacode(row, 1)
            if result:
                unique_gtins.add(result)
    return sorted(unique_gtins)


def is_datacode_valid(string) -> bool:
    encoded_string = string.replace('\\x1d', '\x1d')
    long_pattern = r"^01\d{14}21(?:[\w!\"%&'()*+,\-./_:;=<>?]{6}|[\w!\"%&'()*+,\-./_:;=<>?]{13}|" \
                   r"[\w!\"%&'()*+,\-./_:;=<>?]{20})\x1d91(?:[\w!\"%&'()*+,\-./_:;=<>?]{4})\x1d92" \
                   r"(?:[\w!\"%&'()*+,\-./_:;=<>?]{44}|[\w!\"%&'()*+,\-./_:;=<>?]{88})$"
    short_pattern = r"^01\d{14}21(?:[\w!\"%&'()*+,\-./_:;=<>?]{6}|[\w!\"%&'()*+,\-./_:;=<>?]{7}|" \
                    r"[\w!\"%&'()*+,\-./_:;=<>?]{13})\x1d93(?:[\w!\"%&'()*+,\-./_:;=<>?]{4}|" \
                    r"[\w!\"%&'()*+,\-./_:;=<>?]{7})(\d{10})?$"
    nicotine_pattern = r"^01\d{14}21[\w!\"%&'()*+,\-./_:;=<>?]{7}8005[\w!\"%&'()*+,\-./_:;=<>?]{6}\x1d93" \
                       r"[\w!\"%&'()*+,\-./_:;=<>?]{4}$"
    if (re.match(long_pattern, encoded_string) or re.match(short_pattern, encoded_string) or
            re.match(nicotine_pattern, encoded_string)):
        return True
    else:
        return False


def find_datacode(string, mode=0) -> str:
    """
    :param string:
    :param mode: 0 = matched string, 1 = gtin only, 2 = serial number only
    :return:
    """

    encoded_string = string.replace('\\x1d', '\x1d')

    long_pattern = r"01(?P<gtin>\d{14})21(?P<serial>[\w!\"%&'()*+,\-./_:;=<>?]{6}|[\w!\"%&'()*+,\-./_:;=<>?]{13}|" \
                   r"[\w!\"%&'()*+,\-./_:;=<>?]{20})\x1d91(?P<validation_key>[\w!\"%&'()*+,\-./_:;=<>?]{4})\x1d92" \
                   r"(?P<validation_code>[\w!\"%&'()*+,\-./_:;=<>?]{44}|[\w!\"%&'()*+,\-./_:;=<>?]{88})"

    short_pattern = r"01(?P<gtin>\d{14})21(?P<serial>[\w!\"%&'()*+,\-./_:;=<>?]{6}|[\w!\"%&'()*+,\-./_:;=<>?]{7}|" \
                    r"[\w!\"%&'()*+,\-./_:;=<>?]{13})\x1d93(?P<validation_code>[\w!\"%&'()*+,\-./_:;=<>?]{4}|" \
                    r"[\w!\"%&'()*+,\-./_:;=<>?]{7})(?P<suffix>\d{10})?"

    nicotine_pattern = r"01(?P<gtin>\d{14})21(?P<serial>[\w!\"%&'()*+,\-./_:;=<>?]{7})8005(?P<max_price>" \
                       r"[\w!\"%&'()*+,\-./_:;=<>?]{6})\x1d93(?P<validation_code>[\w!\"%&'()*+,\-./_:;=<>?]{4})"

    long_match = re.search(long_pattern, encoded_string)
    short_match = re.search(short_pattern, encoded_string)
    nicotine_match = re.search(nicotine_pattern, encoded_string)

    if long_match:
        return long_match.group(mode)
    elif short_match:
        return short_match.group(mode)
    elif nicotine_match:
        return nicotine_match.group(mode)
    else:
        return ''


def find_datacode_without_validation_key(string) -> str:
    """
    :param string:
    :param mode: 0 = matched string, 1 = gtin only, 2 = serial number only
    :return:
    """

    encoded_string = string.replace('\\x1d', '\x1d')

    long_pattern = r"01(?P<gtin>\d{14})21(?P<serial>[\w!\"%&'()*+,\-./_:;=<>?]{6}|[\w!\"%&'()*+,\-./_:;=<>?]{13}|" \
                   r"[\w!\"%&'()*+,\-./_:;=<>?]{20})\x1d91(?P<validation_key>[\w!\"%&'()*+,\-./_:;=<>?]{4})\x1d92" \
                   r"(?P<validation_code>[\w!\"%&'()*+,\-./_:;=<>?]{44}|[\w!\"%&'()*+,\-./_:;=<>?]{88})"

    short_pattern = r"01(?P<gtin>\d{14})21(?P<serial>[\w!\"%&'()*+,\-./_:;=<>?]{6}|[\w!\"%&'()*+,\-./_:;=<>?]{7}|" \
                    r"[\w!\"%&'()*+,\-./_:;=<>?]{13})\x1d93(?P<validation_code>[\w!\"%&'()*+,\-./_:;=<>?]{4}|" \
                    r"[\w!\"%&'()*+,\-./_:;=<>?]{7})(?P<suffix>\d{10})?"

    nicotine_pattern = r"01(?P<gtin>\d{14})21(?P<serial>[\w!\"%&'()*+,\-./_:;=<>?]{7})8005(?P<max_price>" \
                       r"[\w!\"%&'()*+,\-./_:;=<>?]{6})\x1d93(?P<validation_code>[\w!\"%&'()*+,\-./_:;=<>?]{4})"

    long_match = re.search(long_pattern, encoded_string)
    short_match = re.search(short_pattern, encoded_string)
    nicotine_match = re.search(nicotine_pattern, encoded_string)

    if long_match:
        return re.sub(r'\x1d91.*', '', long_match.group(0))
    elif short_match:
        return re.sub(r'\x1d93.*', '', short_match.group(0))
    elif nicotine_match:
        return re.sub(r'\x1d93.*', '', nicotine_match.group(0))
    else:
        return ''


def join_strings(csv_file_path='/home/usr/PycharmProjects/matrix_tags/data/test/long_strings.csv',
                 txt_file_path='/home/usr/PycharmProjects/matrix_tags/data/test/short_strings.txt',
                 output_file_path='/home/usr/PycharmProjects/matrix_tags/data/test/output.csv'):
    # Open the CSV file and read the long strings

    with open(csv_file_path) as file:
        rows = file.readlines()
        long_strings = [find_datacode(row) for row in rows if find_datacode(row)]
    # with open(csv_file_path, 'r') as csv_file:
    #     csv_reader = csv.reader(csv_file, delimiter='\t', quoting=csv.QUOTE_NONE, escapechar='\\')
    #     long_strings = [row[0] for row in csv_reader if len(row) > 0 and isinstance(row[0], str) and
    #                     is_datacode_valid(row[0])]

    # Open the TXT file and read the shorter strings and values
    with open(txt_file_path, 'r') as txt_file:
        lines = txt_file.readlines()
        short_strings = []
        values = []
        for line in lines:
            parts = line.strip().split(',')
            short_strings.append(parts[0])
            values.append(tuple(parts[1:]))

    # Join the long strings with the values
    joined_strings = []
    for long_string in long_strings:
        for i, short_string in enumerate(short_strings):
            if short_string in long_string:
                value1, value2, value3, value4, value5, value6, value7, value8, value9 = values[i]
                joined_string = f"{long_string}\t{value1}\t{value2}\t{value3}\t{value4}\t{value5}\t{value6}\t" \
                                f"{value7}\t{value8}\t{value9}"
                joined_strings.append(joined_string)
                break

    # Write the joined strings to the output CSV file
    with open(output_file_path, 'w', newline='') as csv_file:
        csv_writer = csv.writer(csv_file, delimiter='\t', quoting=csv.QUOTE_NONE, escapechar='\\')
        for joined_string in joined_strings:
            csv_writer.writerow(joined_string.split('\t'))


def incorrect_csv_file_codes_count(csv_file_path):
    incorrect_codes_count = 0
    with open(csv_file_path) as file:
        rows = file.readlines()
        for row in rows:
            if not find_datacode(row):
                incorrect_codes_count += 1
    return incorrect_codes_count


def csv_file_row_count(csv_file_path):
    with open(csv_file_path) as file:
        return len(file.readlines())


def is_cis_in_xlsx_file_okay(cis):
    pattern = r"^01(?P<gtin>\d{14})21(?P<serial>[\w!\"%&'()*+,\-./_:;=<>?]{6}|[\w!\"%&'()*+,\-./_:;=<>?]{7}|" \
                    r"[\w!\"%&'()*+,\-./_:;=<>?]{13})$"
    return re.search(pattern, cis) is not None


def fix_xlsx_file_product_cost_value(value):
    if value.find('.') > -1:
        return int(value.replace('.', ''))
    elif value.find(',') > -1:
        return int(value.replace(',', ''))
    else:
        return int(value) * 100


def xlsx_file_exctract_data(xlsx_file_path):
    # TODO: add check for codes in xlsx
    wb = load_workbook(xlsx_file_path)
    ws = wb.active
    result_list = []
    error_count = 0
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3, values_only=True):
        result_list.append({'cis': row[1], 'product_cost': fix_xlsx_file_product_cost_value(row[0]) if row[0] else ''})
        if not is_cis_in_xlsx_file_okay(row[1]):
            error_count += 1
    error_code = 'no_errors' if error_count == 0 else 'errors'
    return result_list, error_code


def csv2xlsx_convert(csv_file_path, xlsx_file_path):
    # TODO: complete this out
    data = pd.read_csv(csv_file_path, sep='\t', lineterminator='\n', header=None)
    data['extracted'] = data[0].apply(find_datacode_without_validation_key)
    data['result'] = data['extracted'].apply(lambda x: x.replace('\x1d', ''))
    data[['result']].to_excel(xlsx_file_path, index=False, header=None)


if __name__ == '__main__':
    file_path = 'data/other_files/test_regex.csv'
    res = gtin_set(file_path)
    # res = incorrect_csv_file_codes_count(file_path)
    print(res)
